import unittest
from datetime import date
from UST import UST, UST_Future
from _datetime import timedelta
import openpyxl

def loadEDSettles(self,stl_date=date(year=date.today().year,month=date.today().month,day=date.today().day)):
        wb=openpyxl.load_workbook('C:/Users/danBackup/eclipse-workspace/sendTEDs/tedSettleCalc2016.xlsm',data_only=True)
        sheet=wb.get_sheet_by_name(str(stl_date.year))
        # dtOfEDSettles is a time delta value in MS Excel format
        dtOfEDSettles=stl_date-date(1899,12,31)
        EDSettles=[]
        for i in range(1,sheet.max_row):
            if sheet.cell(row=i,column=1).internal_value==dtOfEDSettles.days:
                settleRow=i
                break
        for j in range(6,46):
            EDSettles.append(sheet.cell(row=i,column=j).internal_value)
        return EDSettles

def main():
    stlDateZ72s5s=date(2018,1,4)
    stlDateH82s5s=date(2018,4,4)
    stlDateZ710s=date(2017,12,29)
    stlDateH810s=date(2018,3,29)
    b1=UST(date(2019,9,30),stlDateZ72s5s,0.0175,2)
    b2=UST(date(2022,2,28),stlDateZ72s5s,0.01875,2)
    b3=UST(date(2024,8,15),stlDateZ710s,0.02375,2)
    b4=UST(date(2019,12,31),stlDateH82s5s,0.01625,2)
    b5=UST(date(2022,5,31),stlDateH82s5s,0.0175,2)
    b6=UST(date(2024,11,15),stlDateH810s,0.0225,2)
    b7=UST(date(2027,5,15),stlDateZ710s,0.02375,2)
    b8=UST(date(2027,8,15),stlDateH810s,0.0225,2)
    matDates=(date(2018,3,20),date(2018,6,21),date(2018,9,20),date(2018,12,19),
              date(2019,3,19),date(2019,6,20),date(2019,9,19),date(2019,12,18),
              date(2020,3,18),date(2020,6,18),date(2020,9,17),date(2020,12,16),
              date(2021,3,16),date(2021,6,17),date(2021,9,16),date(2021,12,15),
              date(2022,3,15),date(2022,6,16),date(2022,9,15),date(2022,12,21),
              date(2023,3,21),date(2023,6,15),date(2023,9,21),date(2023,12,20),
              date(2024,3,20),date(2024,6,20),date(2024,9,19),date(2024,12,18),
              date(2025,3,18),date(2025,6,19),date(2025,9,18),date(2025,12,17),
              date(2026,3,17),date(2026,6,18),date(2026,9,17),date(2026,12,16),
              date(2027,3,17),date(2027,6,16),date(2027,9,15),date(2027,12,15))
    matRates=loadEDSettles(date(2017,10,17))
    TUZ7=UST_Future(b1,stlDateZ72s5s,date(2017,12,1),1)
    FVZ7=UST_Future(b2,stlDateZ72s5s,date(2017,12,1),1)
    TYZ7=UST_Future(b3,stlDateZ710s,date(2017,12,1),2)
    TUH8=UST_Future(b4,stlDateH82s5s,date(2018,3,1),1)
    FVH8=UST_Future(b5,stlDateH82s5s,date(2018,3,1),1)
    TYH8=UST_Future(b6,stlDateH810s,date(2018,3,1),2)
    TNZ7=UST_Future(b7,stlDateZ710s,date(2017,12,1),2)
    TNH8=UST_Future(b8,stlDateH810s,date(2018,3,1),2)
    
    dlvDt1=date(2017,12,1) # 1st day of delivery month for front contract
    dlvDt2=date(2018,3,1) # 1st day of delivery month for back contract
    print('TUZ7 CF=',TUZ7.CF)
    print('FVZ7 CF=',FVZ7.CF)
    print('TYZ7 CF=',TYZ7.CF)
    print('TNZ7 CF=',TNZ7.CF)
    print('TUH8 CF=',TUH8.CF)
    print('FVH8 CF=',FVH8.CF)
    print('TYH8 CF=',TYH8.CF)
    print('TNH8 CF=',TNH8.CF)
    TUZ7Crv=TUZ7.dfcurve(date(2017,10,17), date(2017,10,18), (100-matRates[0])/100, matDates, matRates)
    print(TUZ7Crv)
    print(TUZ7.bprFromDFCurve(TUZ7Crv, stlDateZ72s5s))
    #print(TUZ7.bimprepo(100.5, 108, date(2017,12,29)))
    #print('b1 previous coupon date',b1.prevCpn(stlDate))
    #print('b1 next coupon date',b1.nextCpn(stlDate))
    #print('b1 number of coupons remaining:',b1.numCoupons('A/A'))
    #print('b1 CF',TUU7.getCF(date(2017,9,1), 1))
    #print('b1 ai',b1.bai(b1.cpn, b1.matDate, b1.stlDate))
    #print('b1 yield',b1.byld(100))
    
    '''
    print('b1 DV01',b1.dv01(0.01424)*10000)
    print('b2 previous coupon date',b2.prevCpn(stlDate))
    print('b2 next coupon date',b2.nextCpn(stlDate))
    print('b2 number of coupons remaining:',b2.numCoupons('A/A'))
    print('b2 CF',FVU7.getCF(date(2017,9,1), 1))
    print('b2 ai',b2.bai(b2.cpn, b2.matDate, b2.stlDate))
    print('b2 yield',b2.byld(100))
    print('b3 previous coupon date',b3.prevCpn(stlDate))
    print('b3 next coupon date',b3.nextCpn(stlDate))
    print('b3 number of coupons remaining:',b3.numCoupons('A/A'))
    print('b3 CF',TYU7.getCF(date(2017,9,1), 2))
    print('b3 yield',b3.byld(100))
    print('b4 previous coupon date',b4.prevCpn(stlDate))
    print('b4 next coupon date',b4.nextCpn(stlDate))
    print('b4 number of coupons remaining:',b4.numCoupons('A/A'))
    print('b4 CF',TUZ7.getCF(date(2017,12,1), 1))
    print('b4 yield',b4.byld(100))
    print('b5 previous coupon date',b5.prevCpn(stlDate))
    print('b5 next coupon date',b5.nextCpn(stlDate))
    print('b5 number of coupons remaining:',b5.numCoupons('A/A'))
    print('b5 CF',FVZ7.getCF(date(2017,12,1), 1))
    print('b5 yield',b5.byld(100))
    print('b6 previous coupon date',b6.prevCpn(stlDate))
    print('b6 next coupon date',b6.nextCpn(stlDate))
    print('b6 number of coupons remaining:',b6.numCoupons('A/A'))
    print('b6 CF',TYZ7.getCF(date(2017,12,1), 2))
    print('b6 yield',b6.byld(100))
    print('b7 CF',TNU7.getCF(date(2017,9,1), 2))
    print('b7 yield',b7.byld(100))
    print('b8 CF',TNZ7.getCF(date(2017,12,1), 2))
    print('b8 yield',b8.byld(100))
    '''
    
if __name__=="__main__":
    main()
